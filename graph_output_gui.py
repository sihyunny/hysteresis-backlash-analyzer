#!/usr/bin/env python3.11
# -*- coding: utf-8 -*-

"""
íˆìŠ¤í…Œë¦¬ì‹œìŠ¤ ê³¡ì„  ë¶„ì„ í”„ë¡œê·¸ë¨ - ìµœì¢… ë²„ì „
ë‹¤ì–‘í•œ ë°ì´í„° í˜•ì‹ì— ëŒ€ì‘ (í° ë³€í™” + ì—°ì†ì  ë³€í™”)
í† í¬ ê°•ì„± ê³„ì‚° ì¶”ê°€
"""

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import sys
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment

def find_hysteresis_loop_v2(x_data, y_data):
    """
    ê°œì„ ëœ íˆìŠ¤í…Œë¦¬ì‹œìŠ¤ ë£¨í”„ ê°ì§€
    ë°©ë²• 1: Xê°’ ìµœëŒ€ ì§€ì  ê¸°ì¤€ (test001.xlsx íƒ€ì…) - ìš°ì„  ì‚¬ìš©
    ë°©ë²• 2: í° Xê°’ ë³€í™” ê°ì§€ (test1.xlsx íƒ€ì…) - ë°©ë²• 1ì´ ë¶€ì í•©í•  ë•Œ ì‚¬ìš©
    """
    # ë°©ë²• 1: Xê°’ ìµœëŒ€ ì§€ì  ê¸°ì¤€ (ê¸°ë³¸ ë°©ë²•)
    x_max_idx = np.argmax(x_data)
    x_max = x_data[x_max_idx]
    
    # Xê°’ì´ ìµœëŒ€ ì§€ì  ê·¼ì²˜ì— ì—¬ëŸ¬ ê°œ ìˆì„ ìˆ˜ ìˆìœ¼ë¯€ë¡œ ì²« ë²ˆì§¸ ìµœëŒ€ê°’ ì°¾ê¸°
    first_max_idx = np.where(x_data == x_max)[0][0]
    
    ascending_start = 0
    descending_start = first_max_idx
    descending_end = len(x_data)
    
    # ë°©ë²• 1ì´ ì í•©í•œì§€ í™•ì¸ (ìƒìŠ¹ ê²½ë¡œì™€ í•˜ê°• ê²½ë¡œê°€ ëª¨ë‘ ì¶©ë¶„í•œì§€)
    if descending_start > 10 and (descending_end - descending_start) > 10:
        return ascending_start, descending_start, descending_end, "X ìµœëŒ€ê°’ ê¸°ì¤€"
    
    # ë°©ë²• 2: í° ë³€í™” ê°ì§€ (ë°©ë²• 1ì´ ë¶€ì í•©í•  ë•Œ)
    x_diff = np.diff(x_data)
    
    large_increases = []
    large_decreases = []
    
    for i in range(len(x_diff)):
        if x_diff[i] > 1.0:
            large_increases.append(i)
        if x_diff[i] < -1.0:
            large_decreases.append(i)
    
    # í° ë³€í™”ê°€ ìˆìœ¼ë©´ ë°©ë²• 2 ì‚¬ìš©
    if len(large_increases) > 0 and len(large_decreases) > 0:
        ascending_start = large_increases[0] + 1
        descending_start = None
        
        for dec_idx in large_decreases:
            if dec_idx > ascending_start:
                descending_start = dec_idx + 1
                break
        
        if descending_start is not None:
            descending_end = len(x_data)
            for inc_idx in large_increases:
                if inc_idx > descending_start:
                    descending_end = inc_idx + 1
                    break
            
            return ascending_start, descending_start, descending_end, "í° ë³€í™” ê°ì§€"
    
    # ë‘˜ ë‹¤ ì‹¤íŒ¨í•˜ë©´ ë°©ë²• 1ì˜ ê²°ê³¼ ë°˜í™˜
    return 0, first_max_idx, len(x_data), "X ìµœëŒ€ê°’ ê¸°ì¤€ (ê¸°ë³¸)"

def calculate_backlash(file_path):
    """íˆìŠ¤í…Œë¦¬ì‹œìŠ¤ ë£¨í”„ë¥¼ ìë™ ê°ì§€í•˜ì—¬ ë°±ë˜ì‰¬ ê³„ì‚°"""
    # íŒŒì¼ í™•ì¥ìì— ë”°ë¼ ì ì ˆí•œ í•¨ìˆ˜ë¡œ ì½ê¸°
    file_ext = os.path.splitext(file_path)[1].lower()
    if file_ext == '.csv':
        df = pd.read_csv(file_path, header=None)
    else:  # .xlsx, .xls ë“±
        df = pd.read_excel(file_path, header=None)

    # Aì—´(0ë²ˆ ì—´)ì´ 0ì¸ í–‰ë“¤ì„ ì œê±°
    df = df[df[0] != 0]

    x_data = df[0].values
    y_data = df[2].values

    # Xì¶• ê¸°ì¤€ê°’ ê³„ì‚°
    x_max = x_data.max()
    x_min = x_data.min()
    
    # Xì¶• 3% ì§€ì ê³¼ -3% ì§€ì  ê³„ì‚°
    x_3_percent = x_max * 0.03
    x_minus_3_percent = x_min * 0.03
    
    # íˆìŠ¤í…Œë¦¬ì‹œìŠ¤ ë£¨í”„ ìë™ ê°ì§€
    ascending_start, descending_start, descending_end, method = find_hysteresis_loop_v2(x_data, y_data)
    
    if ascending_start is None or descending_start is None:
        return None
    
    # ìƒìŠ¹ ê²½ë¡œ
    ascending_x = x_data[ascending_start:descending_start]
    ascending_y = y_data[ascending_start:descending_start]
    
    # í•˜ê°• ê²½ë¡œ
    descending_x = x_data[descending_start:descending_end]
    descending_y = y_data[descending_start:descending_end]
    
    # ì„ í˜• ë³´ê°„ì„ ì‚¬ìš©í•˜ì—¬ íŠ¹ì • Xê°’ì—ì„œì˜ Yê°’ ì°¾ê¸°
    def find_y_at_x(x_arr, y_arr, target_x):
        if len(x_arr) == 0:
            return None
        
        # target_xì— ì •í™•íˆ ì¼ì¹˜í•˜ëŠ” ì ì´ ìˆëŠ”ì§€ í™•ì¸
        exact_match = np.where(np.abs(x_arr - target_x) < 0.01)[0]
        if len(exact_match) > 0:
            return y_arr[exact_match[0]]
        
        # target_xë³´ë‹¤ ì‘ê±°ë‚˜ ê°™ì€ ì ê³¼ í¬ê±°ë‚˜ ê°™ì€ ì  ì°¾ê¸°
        lower_indices = np.where(x_arr <= target_x)[0]
        upper_indices = np.where(x_arr >= target_x)[0]
        
        if len(lower_indices) > 0 and len(upper_indices) > 0:
            lower_idx = lower_indices[np.argmax(x_arr[lower_indices])]
            upper_idx = upper_indices[np.argmin(x_arr[upper_indices])]
            
            x1, y1 = x_arr[lower_idx], y_arr[lower_idx]
            x2, y2 = x_arr[upper_idx], y_arr[upper_idx]
            
            if x2 != x1:
                y_interpolated = y1 + (y2 - y1) * (target_x - x1) / (x2 - x1)
                return y_interpolated
            else:
                return y1
        elif len(lower_indices) > 0:
            return y_arr[lower_indices[-1]]
        elif len(upper_indices) > 0:
            return y_arr[upper_indices[0]]
        
        return None
    
    # back1: í•˜ê°• ê²½ë¡œì—ì„œ Xì¶• 3% ì§€ì ì˜ Yê°’ (ì„ í˜• ë³´ê°„ ì‚¬ìš©)
    back1 = find_y_at_x(descending_x, descending_y, x_3_percent)

    # back2: ì „ì²´ ë°ì´í„°ì—ì„œ Xì¶• -3% ì§€ì ì˜ Yê°’ (ì„ í˜• ë³´ê°„ ì‚¬ìš©)
    # X=-3%ë¥¼ ì§€ë‚˜ê°€ëŠ” ëª¨ë“  êµ¬ê°„ì—ì„œ ì„ í˜• ë³´ê°„í•œ í›„, ê°€ì¥ ì‘ì€ Yê°’ ì„ íƒ
    back2 = None
    back2_candidates = []

    for i in range(len(x_data) - 1):
        x1, x2 = x_data[i], x_data[i + 1]
        y1, y2 = y_data[i], y_data[i + 1]

        # X=-3% ì§€ì ì´ ì´ êµ¬ê°„ [x1, x2] ì‚¬ì´ì— ìˆëŠ”ì§€ í™•ì¸
        if (x1 <= x_minus_3_percent <= x2) or (x2 <= x_minus_3_percent <= x1):
            # ì„ í˜• ë³´ê°„ìœ¼ë¡œ Yê°’ ê³„ì‚°
            if x2 != x1:
                y_interpolated = y1 + (y2 - y1) * (x_minus_3_percent - x1) / (x2 - x1)
                back2_candidates.append(y_interpolated)

    # ëª¨ë“  í›„ë³´ ì¤‘ ê°€ì¥ ì‘ì€ Yê°’ ì„ íƒ
    if len(back2_candidates) > 0:
        back2 = min(back2_candidates)
    else:
        # í›„ë³´ê°€ ì—†ìœ¼ë©´ ê¸°ì¡´ ë°©ì‹ ì‚¬ìš©
        back2 = find_y_at_x(x_data, y_data, x_minus_3_percent)

    # í† í¬ ê°•ì„± ê³„ì‚°ì„ ìœ„í•œ tolerance ë³€ìˆ˜ (ê¸°ì¡´ ìœ„ì¹˜ ìœ ì§€)
    tolerance = abs(x_max - x_min) * 0.01  # 1% í—ˆìš© ì˜¤ì°¨
    
    # ë°±ë˜ì‰¬ ê³„ì‚°
    backlash = None
    if back1 is not None and back2 is not None:
        backlash = abs(back1 - back2)
    
    # í† í¬ ê°•ì„± ê³„ì‚°
    # X=50% ì§€ì ê³¼ X=100% ì§€ì  ê³„ì‚°
    x_50_percent = x_max * 0.5
    x_100_percent = x_max
    
    # X=100% ì§€ì ì˜ Yê°’ (ê°€ì¥ ê°€ê¹Œìš´ ì‹¤ì œ ë°ì´í„° í¬ì¸íŠ¸)
    y_100_percent = find_y_at_x(x_data, y_data, x_100_percent)
    
    # X=50% ì§€ì ì˜ ë‘ Yê°’ ì°¾ê¸° (ìƒìŠ¹ ê²½ë¡œì™€ í•˜ê°• ê²½ë¡œ)
    mask_50 = np.abs(x_data - x_50_percent) <= tolerance
    y_50_percent_min = None
    y_50_percent_all = []
    
    if np.any(mask_50):
        y_50_percent_all = y_data[mask_50].tolist()
        y_50_percent_min = np.min(y_data[mask_50])  # ë” ì‘ì€ Yê°’
    else:
        # í—ˆìš© ì˜¤ì°¨ ë‚´ì— ë°ì´í„°ê°€ ì—†ìœ¼ë©´ ì„ í˜• ë³´ê°„ ì‚¬ìš©
        y_50_percent_min = find_y_at_x(x_data, y_data, x_50_percent)
        if y_50_percent_min is not None:
            y_50_percent_all = [y_50_percent_min]
    
    # ìš°ì¸¡ í† í¬ ê°•ì„± ê³„ì‚°: b/a
    b = x_100_percent - x_50_percent  # Xê°’ì˜ ì°¨ì´
    a = None
    torque_stiffness = None
    
    if y_100_percent is not None and y_50_percent_min is not None:
        a = y_100_percent - y_50_percent_min  # Yê°’ì˜ ì°¨ì´
        if a != 0:
            torque_stiffness = b / a
    
    # ì¢Œì¸¡ í† í¬ ê°•ì„± ê³„ì‚° (X=-50%, X=-100%)
    x_minus_50_percent = x_min * 0.5
    x_minus_100_percent = x_min
    
    # X=-100% ì§€ì ì˜ Yê°’
    y_minus_100_percent = find_y_at_x(x_data, y_data, x_minus_100_percent)
    
    # X=-50% ì§€ì ì˜ ë‘ Yê°’ ì°¾ê¸° (ìƒìŠ¹ ê²½ë¡œì™€ í•˜ê°• ê²½ë¡œ)
    mask_minus_50 = np.abs(x_data - x_minus_50_percent) <= tolerance
    y_minus_50_percent_max = None
    y_minus_50_percent_all = []
    
    if np.any(mask_minus_50):
        y_minus_50_percent_all = y_data[mask_minus_50].tolist()
        y_minus_50_percent_max = np.max(y_data[mask_minus_50])  # ë” í° Yê°’
    else:
        # í—ˆìš© ì˜¤ì°¨ ë‚´ì— ë°ì´í„°ê°€ ì—†ìœ¼ë©´ ì„ í˜• ë³´ê°„ ì‚¬ìš©
        y_minus_50_percent_max = find_y_at_x(x_data, y_data, x_minus_50_percent)
        if y_minus_50_percent_max is not None:
            y_minus_50_percent_all = [y_minus_50_percent_max]
    
    # ì¢Œì¸¡ í† í¬ ê°•ì„± ê³„ì‚°: b_left/a_left
    b_left = x_minus_50_percent - x_minus_100_percent  # Xê°’ì˜ ì°¨ì´ (ì ˆëŒ€ê°’)
    a_left = None
    torque_stiffness_left = None
    
    if y_minus_100_percent is not None and y_minus_50_percent_max is not None:
        a_left = y_minus_100_percent - y_minus_50_percent_max  # Yê°’ì˜ ì°¨ì´
        if a_left != 0:
            torque_stiffness_left = abs(b_left / a_left)
    
    return {
        'x_data': x_data,
        'y_data': y_data,
        'ascending_x': ascending_x,
        'ascending_y': ascending_y,
        'descending_x': descending_x,
        'descending_y': descending_y,
        'x_max': x_max,
        'x_min': x_min,
        'x_3_percent': x_3_percent,
        'x_minus_3_percent': x_minus_3_percent,
        'back1': back1,
        'back2': back2,
        'backlash': backlash,
        'method': method,
        'x_50_percent': x_50_percent,
        'x_100_percent': x_100_percent,
        'y_50_percent_min': y_50_percent_min,
        'y_50_percent_all': y_50_percent_all,
        'y_100_percent': y_100_percent,
        'b': b,
        'a': a,
        'torque_stiffness': torque_stiffness,
        'x_minus_50_percent': x_minus_50_percent,
        'x_minus_100_percent': x_minus_100_percent,
        'y_minus_50_percent_max': y_minus_50_percent_max,
        'y_minus_50_percent_all': y_minus_50_percent_all,
        'y_minus_100_percent': y_minus_100_percent,
        'b_left': b_left,
        'a_left': a_left,
        'torque_stiffness_left': torque_stiffness_left
    }

def plot_and_save(result, output_path, input_filename=None):
    """ê·¸ë˜í”„ ê·¸ë¦¬ê¸° ë° ì €ì¥"""
    if not result:
        print("ë¶„ì„ ê²°ê³¼ê°€ ì—†ìŠµë‹ˆë‹¤.")
        return False
    
    ascending_x = result['ascending_x']
    ascending_y = result['ascending_y']
    descending_x = result['descending_x']
    descending_y = result['descending_y']
    x_3_percent = result['x_3_percent']
    x_minus_3_percent = result['x_minus_3_percent']
    back1 = result['back1']
    back2 = result['back2']
    backlash = result['backlash']
    
    # ìš°ì¸¡ í† í¬ ê°•ì„± ê´€ë ¨ ë³€ìˆ˜
    x_50_percent = result.get('x_50_percent')
    x_100_percent = result.get('x_100_percent')
    y_50_percent_min = result.get('y_50_percent_min')
    y_100_percent = result.get('y_100_percent')
    b = result.get('b')
    a = result.get('a')
    torque_stiffness = result.get('torque_stiffness')
    
    # ì¢Œì¸¡ í† í¬ ê°•ì„± ê´€ë ¨ ë³€ìˆ˜
    x_minus_50_percent = result.get('x_minus_50_percent')
    x_minus_100_percent = result.get('x_minus_100_percent')
    y_minus_50_percent_max = result.get('y_minus_50_percent_max')
    y_minus_100_percent = result.get('y_minus_100_percent')
    b_left = result.get('b_left')
    a_left = result.get('a_left')
    torque_stiffness_left = result.get('torque_stiffness_left')
    
    # ê·¸ë˜í”„ ìƒì„± (ë¡œê³  ê³µê°„ ì¶”ê°€)
    import matplotlib.gridspec as gridspec
    from PIL import Image

    # Figure ìƒì„± (ê°€ë¡œë¡œ ê¸´ ë¹„ìœ¨ - ë¡œê³  ê³µê°„ í¬í•¨)
    fig = plt.figure(figsize=(12, 8.5))

    # GridSpecìœ¼ë¡œ ë ˆì´ì•„ì›ƒ êµ¬ì„±: ìƒë‹¨ ë¡œê³ , í•˜ë‹¨ ê·¸ë˜í”„
    graph_left = 0.08  # ê·¸ë˜í”„ ì™¼ìª½ ì—¬ë°±
    gs = gridspec.GridSpec(2, 1, height_ratios=[0.8, 9], hspace=0.02,
                           left=graph_left, right=0.98,
                           top=0.98, bottom=0.06)

    # ìƒë‹¨: ë¡œê³  ì˜ì—­
    ax_logo = fig.add_subplot(gs[0])
    ax_logo.axis('off')

    # ë¡œê³  íŒŒì¼ ë¡œë“œ ë° í‘œì‹œ
    script_dir = os.path.dirname(os.path.abspath(__file__))
    logo_path = os.path.join(script_dir, 'logo.png')
    if os.path.exists(logo_path):
        try:
            logo_img = Image.open(logo_path)
            # ë¡œê³  í¬ê¸° ê³„ì‚° (ì›ë³¸ì˜ ì ˆë°˜ í¬ê¸°, ë¹„ìœ¨ ìœ ì§€)
            img_width = logo_img.width / 3.0
            img_height = logo_img.height / 3.0

            # ì¤‘ì•™ ì •ë ¬ì„ ìœ„í•œ extent ê³„ì‚°
            # ì „ì²´ ì˜ì—­ì˜ ì¤‘ì•™ì„ ê¸°ì¤€ìœ¼ë¡œ ë°°ì¹˜
            center_x = logo_img.width / 2.0
            left = center_x - img_width / 2.0
            right = center_x + img_width / 2.0

            # ë¡œê³ ë¥¼ ì¤‘ì•™ ìƒë‹¨ì— ë°°ì¹˜
            ax_logo.imshow(logo_img, extent=[left, right, img_height, 0])
            ax_logo.set_xlim(0, logo_img.width)
            ax_logo.set_ylim(logo_img.height, 0)
        except Exception as e:
            print(f"ë¡œê³  ë¡œë“œ ì‹¤íŒ¨: {e}")

    # í•˜ë‹¨: ë©”ì¸ ê·¸ë˜í”„ ì˜ì—­
    ax_main = fig.add_subplot(gs[1])
    plt.sca(ax_main)  # í˜„ì¬ axesë¥¼ ë©”ì¸ ê·¸ë˜í”„ë¡œ ì„¤ì •

    # ì „ì²´ íˆìŠ¤í…Œë¦¬ì‹œìŠ¤ ê³¡ì„ ì„ íŒŒë€ìƒ‰ìœ¼ë¡œ í‘œì‹œ
    x_data = result['x_data']
    y_data = result['y_data']
    plt.plot(x_data, y_data, 'b-', linewidth=1.2, alpha=0.8, label='Hysteresis Curve')
    plt.scatter(x_data, y_data, c='orange', s=10, alpha=1.0, zorder=3, label='Data Points')
    
    # Xì¶• 3% ì§€ì ê³¼ -3% ì§€ì  í‘œì‹œ
    if back1 is not None:
        plt.scatter([x_3_percent], [back1], c='red', s=100, marker='o',
                    edgecolors='black', linewidths=2,
                    label=f'X=3%: Y={back1:.4f}', zorder=5)
        plt.axvline(x=x_3_percent, color='red', linestyle='--', linewidth=1.5, alpha=0.6)
        # 3% í…ìŠ¤íŠ¸ í‘œì‹œ
        plt.text(x_3_percent, result['y_data'].min() * 0.95, '3%',
                fontsize=10, fontweight='bold', color='red',
                horizontalalignment='center', verticalalignment='top')

    if back2 is not None:
        plt.scatter([x_minus_3_percent], [back2], c='green', s=100, marker='s',
                    edgecolors='black', linewidths=2,
                    label=f'X=-3%: Y={back2:.4f}', zorder=5)
        plt.axvline(x=x_minus_3_percent, color='green', linestyle='--', linewidth=1.5, alpha=0.6)
        # -3% í…ìŠ¤íŠ¸ í‘œì‹œ
        plt.text(x_minus_3_percent, result['y_data'].min() * 0.95, '-3%',
                fontsize=10, fontweight='bold', color='green',
                horizontalalignment='center', verticalalignment='top')
    
    # ë°±ë˜ì‰¬ë¥¼ ì‹œê°ì ìœ¼ë¡œ í‘œì‹œ (ìˆ˜í‰ì„  + ì–‘ë°©í–¥ í™”ì‚´í‘œ)
    if back1 is not None and back2 is not None:
        # back1ì—ì„œ ìˆ˜í‰ì„  (ì˜¤ë¥¸ìª½ìœ¼ë¡œ)
        x_line_end = result['x_data'].max() * 0.2  # ê·¸ë˜í”„ ì¤‘ê°„ê¹Œì§€
        plt.plot([x_3_percent, x_line_end], [back1, back1], 
                'red', linewidth=1.5, linestyle='-', alpha=0.7, zorder=4)
        
        # back2ì—ì„œ ìˆ˜í‰ì„  (ì˜¤ë¥¸ìª½ìœ¼ë¡œ)
        plt.plot([x_minus_3_percent, x_line_end], [back2, back2], 
                'green', linewidth=1.5, linestyle='-', alpha=0.7, zorder=4)
        
        # ì–‘ë°©í–¥ í™”ì‚´í‘œ (ìˆ˜í‰ì„  ëì—ì„œ)
        arrow_x = x_line_end * 0.9  # í™”ì‚´í‘œ ìœ„ì¹˜
        plt.annotate('', xy=(arrow_x, back1), xytext=(arrow_x, back2),
                    arrowprops=dict(arrowstyle='<->', color='gray', lw=2, 
                                  mutation_scale=25, shrinkA=0, shrinkB=0),
                    zorder=6)
        
        # ë°±ë˜ì‰¬ í…ìŠ¤íŠ¸ (í™”ì‚´í‘œ ì˜†)
        text_x = arrow_x * 0.6
        text_y = (back1 + back2) / 2
        plt.text(text_x, text_y, f'Back Lash\n={backlash:.2f} arc Min',
                fontsize=14, fontweight='bold',
                verticalalignment='center',
                bbox=dict(boxstyle='round', facecolor='yellow', alpha=0.9, 
                          edgecolor='black', linewidth=2.5),
                zorder=7)
    
    # í† í¬ ê°•ì„± ì‹œê°í™”
    if x_50_percent is not None and x_100_percent is not None:
        # X=50% ì§€ì ì— íŒŒë€ìƒ‰ ìˆ˜ì§ ì ì„  (Y=0 ì´ìƒì—ì„œë§Œ)
        y_max = y_50_percent_min
        plt.plot([x_50_percent, x_50_percent], [0, y_max], color='blue', linestyle='--', linewidth=1.5, alpha=0.6, zorder=2)
        # 50% í…ìŠ¤íŠ¸ í‘œì‹œ
        plt.text(x_50_percent, -0.2, '50%', 
                fontsize=10, fontweight='bold', color='blue',
                horizontalalignment='center', verticalalignment='top')
        
        # X=100% ì§€ì ì— íŒŒë€ìƒ‰ ìˆ˜ì§ ì ì„  (Y=0 ì´ìƒì—ì„œë§Œ)
        plt.plot([x_100_percent, x_100_percent], [0, y_100_percent], color='blue', linestyle='--', linewidth=1.5, alpha=0.6,  zorder=2)
        # 100% í…ìŠ¤íŠ¸ í‘œì‹œ
        plt.text(x_100_percent, -0.2, '100%', 
                fontsize=10, fontweight='bold', color='blue',
                horizontalalignment='center', verticalalignment='top')
        
        if y_50_percent_min is not None and y_100_percent is not None:

            # b ê°’ í‘œì‹œ (Xì¶• ë°©í–¥, í•˜ëŠ˜ìƒ‰)
            mid_y = y_50_percent_min
            plt.annotate('', xy=(x_100_percent, mid_y), xytext=(x_50_percent, mid_y),
                        arrowprops=dict(arrowstyle='<->', color='black', lw=1.5, 
                                      mutation_scale=20, shrinkA=0, shrinkB=0),
                        zorder=6)
            plt.text((x_50_percent + x_100_percent) / 2, mid_y * 1.02, f'b={b:.2f}',
                    fontsize=10, fontweight='bold', color='black',
                    horizontalalignment='center')
            
            # a ê°’ í‘œì‹œ (Yì¶• ë°©í–¥, ì£¼í™©ìƒ‰)
            mid_x = x_100_percent
            plt.annotate('', xy=(mid_x, y_100_percent), xytext=(mid_x, y_50_percent_min),
                        arrowprops=dict(arrowstyle='<->', color='black', lw=1.5, 
                                      mutation_scale=20, shrinkA=0, shrinkB=0),
                        zorder=6)
            plt.text(mid_x * 1.02, (y_50_percent_min + y_100_percent) / 2, f'a={a:.2f}',
                    fontsize=10, fontweight='bold', color='black',
                    verticalalignment='center')
            
            # ìš°ì¸¡ í† í¬ ê°•ì„± í…ìŠ¤íŠ¸ í‘œì‹œ
            if torque_stiffness is not None:
                plt.text(x_100_percent*0.98, mid_y * 0.9, 
                        f'Torsional Stiffness = b/a\n= {torque_stiffness:.4f}',
                        #transform=plt.gca().transAxes,
                        fontsize=12, fontweight='bold',
                        horizontalalignment='right',
                        verticalalignment='top', color='blue')
    
    # ì¢Œì¸¡ í† í¬ ê°•ì„± ì‹œê°í™”
    if x_minus_50_percent is not None and x_minus_100_percent is not None:
        # X=-50% ì§€ì ì— íŒŒë€ìƒ‰ ìˆ˜ì§ ì ì„  (Y=0 ì´í•˜ì—ì„œë§Œ)
        y_min = result['y_data'].min()
        plt.plot([x_minus_50_percent, x_minus_50_percent], [y_minus_50_percent_max, 0], color='blue', linestyle='--', linewidth=1.5, alpha=0.6, zorder=2)
        # -50% í…ìŠ¤íŠ¸ í‘œì‹œ
        plt.text(x_minus_50_percent, -0.2, '50%', 
                fontsize=10, fontweight='bold', color='blue',
                horizontalalignment='center', verticalalignment='top')
        
        # X=-100% ì§€ì ì— íŒŒë€ìƒ‰ ìˆ˜ì§ ì ì„  (Y=0 ì´í•˜ì—ì„œë§Œ)
        plt.plot([x_minus_100_percent, x_minus_100_percent], [y_min, 0], color='blue', linestyle='--', linewidth=1.5, alpha=0.6,  zorder=2)
        # -100% í…ìŠ¤íŠ¸ í‘œì‹œ
        plt.text(x_minus_100_percent, -0.2, '100%', 
                fontsize=10, fontweight='bold', color='blue',
                horizontalalignment='center', verticalalignment='top')
        
        if y_minus_50_percent_max is not None and y_minus_100_percent is not None:
            # b_left ê°’ í‘œì‹œ (Xì¶• ë°©í–¥)
            mid_y_left = y_minus_50_percent_max
            plt.annotate('', xy=(x_minus_50_percent, mid_y_left), xytext=(x_minus_100_percent, mid_y_left),
                        arrowprops=dict(arrowstyle='<->', color='black', lw=1.5, 
                                      mutation_scale=20, shrinkA=0, shrinkB=0),
                        zorder=6)
            plt.text((x_minus_50_percent + x_minus_100_percent) / 2, mid_y_left*1.06 , f'b={abs(b_left):.2f}',
                    fontsize=10, fontweight='bold', color='black',
                    horizontalalignment='center')
            
            # a_left ê°’ í‘œì‹œ (Yì¶• ë°©í–¥)
            mid_x_left = x_minus_100_percent
            plt.annotate('', xy=(mid_x_left, y_minus_100_percent), xytext=(mid_x_left, y_minus_50_percent_max),
                        arrowprops=dict(arrowstyle='<->', color='black', lw=1.5, 
                                      mutation_scale=20, shrinkA=0, shrinkB=0),
                        zorder=6)
            plt.text(mid_x_left * 1, (y_minus_50_percent_max + y_minus_100_percent) / 2, f'a={abs(a_left):.2f}',
                    fontsize=10, fontweight='bold', color='black',
                    verticalalignment='center', horizontalalignment='right')
            
            # ì¢Œì¸¡ í† í¬ ê°•ì„± í…ìŠ¤íŠ¸ í‘œì‹œ
            if torque_stiffness_left is not None:
                plt.text(mid_x_left*0.98,y_minus_50_percent_max*0.9, f'Torsional Stiffness = b/a\n= {torque_stiffness_left:.4f}',
                        fontsize=12, fontweight='bold',
                        horizontalalignment='left',
                        verticalalignment='bottom',color='blue')
    
    plt.xlabel('Torque (Nm)', fontsize=14, fontweight='bold')
    plt.ylabel('Angle (arc min)', fontsize=14, fontweight='bold')
    plt.title('Hysteresis Curve', fontsize=16, fontweight='bold')
    
    # ê²©ìì„  ì„¤ì •
    from matplotlib.ticker import MultipleLocator
    ax = plt.gca()
    
    # Xì¶• ê²©ìì„ : ì£¼ 10 ë‹¨ìœ„, ë³´ì¡° 2.5 ë‹¨ìœ„
    ax.xaxis.set_major_locator(MultipleLocator(10))
    ax.xaxis.set_minor_locator(MultipleLocator(2.5))
    
    # Yì¶• ê²©ìì„ : ì£¼ 10 ë‹¨ìœ„, ë³´ì¡° 2.5 ë‹¨ìœ„
    ax.yaxis.set_major_locator(MultipleLocator(5))
    ax.yaxis.set_minor_locator(MultipleLocator(2.5))
    
    # ì£¼ ê²©ìì„  (Major grid) - íšŒìƒ‰, ì—°í•˜ê³  ì–˜ê²Œ
    plt.grid(True, which='major', alpha=0.5, linestyle='-', linewidth=0.6, color='gray')
    
    # ë³´ì¡° ê²©ìì„  (Minor grid) - ì—°í•œ íšŒìƒ‰, ë” ì—°í•˜ê³  ì–˜ê²Œ
    plt.grid(True, which='minor', alpha=0.25, linestyle='-', linewidth=0.3, color='lightgray')
    
    plt.legend(fontsize=11, loc='best')
    plt.axhline(y=0, color='k', linestyle='-', linewidth=2, alpha=1)
    plt.axvline(x=0, color='k', linestyle='-', linewidth=2, alpha=1)

    # Yì¶•ì„ 0 ì¤‘ì‹¬ìœ¼ë¡œ ëŒ€ì¹­ ì„¤ì •
    y_max_abs = max(abs(result['y_data'].min()), abs(result['y_data'].max()))
    plt.ylim(-y_max_abs * 1.05, y_max_abs * 1.05)  # 5% ì—¬ìœ  ê³µê°„

    # ì—‘ì…€ íŒŒì¼ëª…ì„ ê·¸ë˜í”„ ì™¼ìª½ ëê³¼ ì •ë ¬í•˜ì—¬ í‘œì‹œ
    if input_filename is not None:
        fig.text(graph_left, 0.99, f'File: {input_filename}',
                 fontsize=10, fontweight='bold',
                 verticalalignment='top',
                 horizontalalignment='left',
                 bbox=dict(boxstyle='round', facecolor='green', alpha=0.5,
                          edgecolor='gray', linewidth=1.5))

    # ìƒì„± ë‚ ì§œ/ì‹œê°„ì„ ìš°ì¸¡ ìƒë‹¨ì— í‘œì‹œ (ë””ì§€í„¸ ì‹œê³„ ìŠ¤íƒ€ì¼)
    creation_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    fig.text(0.98, 0.99, creation_datetime,
             fontsize=11, fontweight='bold',
             fontfamily='monospace',
             color='#00008B',  # ì§„í•œ íŒŒë€ìƒ‰ (darkblue)
             verticalalignment='top',
             horizontalalignment='right')

    # tight_layoutì€ GridSpec ì‚¬ìš© ì‹œ ì œê±° (ì´ë¯¸ ë ˆì´ì•„ì›ƒ ì„¤ì •ë¨)
    plt.savefig(output_path, dpi=300, bbox_inches='tight')
    plt.close()
    
    return True

def log_backlash(input_filename, backlash_value):
    """ë°±ë˜ì‰¬ ê°’ì„ ì—‘ì…€ ë¡œê·¸ íŒŒì¼ì— ê¸°ë¡ (ì„œì‹ ì ìš©)"""
    result_dir = 'Result'
    log_file = os.path.join(result_dir, 'backlash_log.xlsx')

    # í˜„ì¬ ë‚ ì§œ/ì‹œê°„
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # ì—‘ì…€ íŒŒì¼ì´ ì—†ìœ¼ë©´ ìƒˆë¡œ ìƒì„±
    if not os.path.exists(log_file):
        wb = Workbook()
        ws = wb.active
        ws.title = "Backlash Log"

        # í—¤ë” ì‘ì„±
        headers = ["DateTime", "FileName", "Backlash(arc min)"]
        ws.append(headers)

        # í—¤ë” ì„œì‹ ì ìš© (1í–‰)
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = Font(bold=True, color="0000FF")  # ë³¼ë“œ, íŒŒë€ìƒ‰
            cell.alignment = Alignment(horizontal="center", vertical="center")

        wb.save(log_file)

    # ê¸°ì¡´ íŒŒì¼ì— ë°ì´í„° ì¶”ê°€
    wb = load_workbook(log_file)
    ws = wb.active

    # ìƒˆ í–‰ ì¶”ê°€
    new_row = [timestamp, input_filename, f"{backlash_value:.4f}"]
    ws.append(new_row)

    # ìƒˆë¡œ ì¶”ê°€ëœ í–‰ì˜ ì…€ì— ì¤‘ì•™ ì •ë ¬ ì ìš©
    row_num = ws.max_row
    for col_num in range(1, 4):
        cell = ws.cell(row=row_num, column=col_num)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    wb.save(log_file)
    print(f"ë°±ë˜ì‰¬ ë¡œê·¸ ê¸°ë¡ ì™„ë£Œ: {log_file}")

def process_file(input_file):
    """íŒŒì¼ ë¶„ì„ ì²˜ë¦¬ í•¨ìˆ˜"""
    
    if not os.path.exists(input_file):
        print(f"ì˜¤ë¥˜: íŒŒì¼ì„ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤: {input_file}")
        sys.exit(1)
    
    # ì—‘ì…€ íŒŒì¼ëª… ì¶”ì¶œ
    input_filename = os.path.basename(input_file)
    
    # Result í´ë” ìƒì„±
    result_dir = 'Result'
    if not os.path.exists(result_dir):
        os.makedirs(result_dir)
        print(f"Result í´ë”ë¥¼ ìƒì„±í–ˆìŠµë‹ˆë‹¤: {result_dir}")
    
    # ì¶œë ¥ íŒŒì¼ ê²½ë¡œ ì„¤ì •
    if len(sys.argv) >= 3:
        output_file = os.path.join(result_dir, os.path.basename(sys.argv[2]))
    else:
        base_name = os.path.splitext(input_filename)[0]
        datetime_str = datetime.now().strftime("%Y.%m.%d_%Hh%Mm%Ss")
        output_file = os.path.join(result_dir, f"{datetime_str}-{base_name}.png")
    
    print(f"ì…ë ¥ íŒŒì¼: {input_file}")
    print(f"ì¶œë ¥ íŒŒì¼: {output_file}")
    print("\në¶„ì„ ì¤‘...")
    
    try:
        # ë¶„ì„ ìˆ˜í–‰
        result = calculate_backlash(input_file)
        
        if result:
            # ê²°ê³¼ ì¶œë ¥
            print("\n========== íˆìŠ¤í…Œë¦¬ì‹œìŠ¤ ê³¡ì„  ë¶„ì„ ê²°ê³¼ ==========")
            print(f"ê°ì§€ ë°©ë²•: {result['method']}")
            print(f"ë°ì´í„° í¬ì¸íŠ¸ ìˆ˜: {len(result['x_data'])}")
            print(f"Xì¶• ë²”ìœ„: {result['x_data'].min():.2f} ~ {result['x_data'].max():.2f}")
            print(f"Yì¶• ë²”ìœ„: {result['y_data'].min():.2f} ~ {result['y_data'].max():.2f}")
            print(f"\nìƒìŠ¹ ê²½ë¡œ: {len(result['ascending_x'])} í¬ì¸íŠ¸")
            print(f"í•˜ê°• ê²½ë¡œ: {len(result['descending_x'])} í¬ì¸íŠ¸")
            print(f"\nXì¶• ìµœëŒ€ê°’ (100%): {result['x_max']:.4f}")
            print(f"Xì¶• ìµœì†Œê°’ (-100%): {result['x_min']:.4f}")
            print(f"\nXì¶• 3% ì§€ì : {result['x_3_percent']:.4f}")
            print(f"Xì¶• -3% ì§€ì : {result['x_minus_3_percent']:.4f}")
            print(f"\nX=3% ì§€ì ì˜ Yê°’: {result['back1']:.4f}")
            print(f"X=-3% ì§€ì ì˜ Yê°’: {result['back2']:.4f}")
            print(f"\në°±ë˜ì‰¬ (Backlash): {result['backlash']:.4f}")
            
            # ìš°ì¸¡ í† í¬ ê°•ì„± ê²°ê³¼ ì¶œë ¥
            if result.get('torque_stiffness') is not None:
                print(f"\n========== ìš°ì¸¡ í† í¬ ê°•ì„± ê³„ì‚° ê²°ê³¼ ==========")
                print(f"X=50% ì§€ì : {result['x_50_percent']:.4f}")
                print(f"X=100% ì§€ì : {result['x_100_percent']:.4f}")
                print(f"Y (X=50%, ìµœì†Œê°’): {result['y_50_percent_min']:.4f}")
                print(f"Y (X=100%): {result['y_100_percent']:.4f}")
                print(f"b (Xê°’ ì°¨ì´): {result['b']:.4f}")
                print(f"a (Yê°’ ì°¨ì´): {result['a']:.4f}")
                print(f"í† í¬ ê°•ì„± (b/a): {result['torque_stiffness']:.4f}")
            
            # ì¢Œì¸¡ í† í¬ ê°•ì„± ê²°ê³¼ ì¶œë ¥
            if result.get('torque_stiffness_left') is not None:
                print(f"\n========== ì¢Œì¸¡ í† í¬ ê°•ì„± ê³„ì‚° ê²°ê³¼ ==========")
                print(f"X=-50% ì§€ì : {result['x_minus_50_percent']:.4f}")
                print(f"X=-100% ì§€ì : {result['x_minus_100_percent']:.4f}")
                print(f"Y (X=-50%, ìµœëŒ€ê°’): {result['y_minus_50_percent_max']:.4f}")
                print(f"Y (X=-100%): {result['y_minus_100_percent']:.4f}")
                print(f"b_left (Xê°’ ì°¨ì´): {abs(result['b_left']):.4f}")
                print(f"a_left (Yê°’ ì°¨ì´): {abs(result['a_left']):.4f}")
                print(f"í† í¬ ê°•ì„± (b/a): {result['torque_stiffness_left']:.4f}")
            
            print("=" * 50)
            
            # ê·¸ë˜í”„ ì €ì¥
            if plot_and_save(result, output_file, input_filename):
                print(f"\nê·¸ë˜í”„ê°€ ì„±ê³µì ìœ¼ë¡œ ì €ì¥ë˜ì—ˆìŠµë‹ˆë‹¤: {output_file}")

                # ë°±ë˜ì‰¬ ê°’ ë¡œê·¸ ê¸°ë¡
                log_backlash(input_filename, result['backlash'])

                return output_file  # ì¶œë ¥ íŒŒì¼ ê²½ë¡œ ë°˜í™˜
            else:
                print("\nê·¸ë˜í”„ ì €ì¥ì— ì‹¤íŒ¨í–ˆìŠµë‹ˆë‹¤.")
                sys.exit(1)
        else:
            print("ë¶„ì„ì— ì‹¤íŒ¨í–ˆìŠµë‹ˆë‹¤.")
            sys.exit(1)

    except Exception as e:
        print(f"\nì˜¤ë¥˜ ë°œìƒ: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

def main_cli():
    """ì»¤ë§¨ë“œë¼ì¸ ëª¨ë“œ"""
    if len(sys.argv) < 2:
        print("ì‚¬ìš©ë²•: python3.11 graph_output.py <ë°ì´í„°íŒŒì¼ê²½ë¡œ(.xlsx/.xls/.csv)> [ì¶œë ¥íŒŒì¼ê²½ë¡œ]")
        sys.exit(1)
    
    input_file = sys.argv[1]
    process_file(input_file)

def main_gui():
    """ê°„ë‹¨í•œ GUI ëª¨ë“œ - CustomTkinter"""
    import customtkinter as ctk
    from tkinter import filedialog, messagebox
    import threading
    from PIL import Image

    # CustomTkinter í…Œë§ˆ ì„¤ì •
    ctk.set_appearance_mode("dark")
    ctk.set_default_color_theme("blue")
    
    class HysteresisApp:
        def __init__(self, root):
            self.root = root
            self.root.title("Backlash Analyzer")
            self.root.geometry("750x650")
            self.root.resizable(False, False)

            # ì„ íƒëœ íŒŒì¼ ê²½ë¡œ
            self.selected_file = None
            # ì €ì¥ëœ ì¶œë ¥ íŒŒì¼ ê²½ë¡œ
            self.output_file = None

            # UI êµ¬ì„±
            self.create_widgets()
            
        def create_widgets(self):
            # ì œëª© ë° ë¡œê³  í”„ë ˆì„
            header_frame = ctk.CTkFrame(self.root, fg_color="transparent")
            header_frame.pack(pady=15, padx=20, fill="x")

            # ì œëª© (ì¢Œì¸¡)
            title_label = ctk.CTkLabel(header_frame, text="Backlash Analyzer",
                                       font=("Consolas", 36, "bold"))
            title_label.pack(side="left")

            # ë¡œê³  (ìš°ì¸¡)
            try:
                logo_image = Image.open("logo.png")
                # ë¡œê³  í¬ê¸° ì¡°ì • (ë†’ì´ 30px ì •ë„ë¡œ)
                logo_height = 30
                aspect_ratio = logo_image.width / logo_image.height
                logo_width = int(logo_height * aspect_ratio)

                logo_ctk = ctk.CTkImage(light_image=logo_image,
                                       dark_image=logo_image,
                                       size=(logo_width, logo_height))
                logo_label = ctk.CTkLabel(header_frame, image=logo_ctk, text="")
                logo_label.pack(side="right")
            except Exception as e:
                print(f"ë¡œê³  ë¡œë“œ ì‹¤íŒ¨: {e}")

            # ì‚¬ìš© ê°€ì´ë“œ í”„ë ˆì„
            guide_frame = ctk.CTkFrame(self.root)
            guide_frame.pack(pady=10, padx=0, fill="x")

            guide_title = ctk.CTkLabel(guide_frame, text="ğŸ“¢ ì‚¬ìš© ê°€ì´ë“œ",
                                       font=("Consolas", 21, "bold"))
            guide_title.pack(pady=(10, 5))

            guide_text = (
                "â€¢ íŒŒì¼ í˜•ì‹: Excel (.xlsx, .xls) ë˜ëŠ” CSV (.csv)\n\n"
                "â€¢ ë°ì´í„° í˜•ì‹: 1ì—´(Aì—´) = í† í¬(Nm), 3ì—´(Cì—´) = ê°ë„(arc min)\n\n"
                "â€¢ í—¤ë” ì—†ì´ ë°ì´í„°ë§Œ í¬í•¨ë˜ì–´ì•¼ í•©ë‹ˆë‹¤ (1ì—´ = 0ì¸ í–‰ì€ ì œì™¸)\n\n"
                "â€¢ ë¶„ì„ ê²°ê³¼: ë°±ë˜ì‰¬(Backlash), í† í¬ ê°•ì„±(Torsional Stiffness)\n\n"
                "â€¢ ì¶œë ¥: Result í´ë”ì— PNG ê·¸ë˜í”„ + backlash_log.xlsx ì €ì¥"
            )

            guide_label = ctk.CTkLabel(guide_frame, text=guide_text,
                                       font=("Consolas", 18),
                                       justify="left")
            guide_label.pack(anchor="w", padx=15, pady=(0, 10))

            # íŒŒì¼ ì„ íƒ í”„ë ˆì„
            file_frame = ctk.CTkFrame(self.root)
            file_frame.pack(pady=12, padx=20, fill="x")

            file_label = ctk.CTkLabel(file_frame, text="ë°ì´í„° íŒŒì¼:",
                                     font=("Consolas", 20, "bold"))
            file_label.pack(side="left", padx=(10, 5))

            browse_btn = ctk.CTkButton(file_frame, text="íŒŒì¼ ì„ íƒ",
                                      command=self.browse_file,
                                      font=("Consolas", 18, "bold"),
                                      fg_color="#4CAF50",
                                      hover_color="#45a049",
                                      width=130, height=50)
            browse_btn.pack(side="right", padx=(5, 10))

            self.file_entry = ctk.CTkEntry(file_frame, font=("Consolas", 18),
                                          height=50)
            self.file_entry.pack(side="left", padx=5, fill="x", expand=True)

            # ë²„íŠ¼ í”„ë ˆì„ (ë¶„ì„ ì‹œì‘ + Result í´ë” ì—´ê¸°)
            button_frame = ctk.CTkFrame(self.root, fg_color="transparent")
            button_frame.pack(pady=12, padx=20, fill="x")

            # ë¶„ì„ ë²„íŠ¼
            self.analyze_btn = ctk.CTkButton(button_frame, text="ë¶„ì„ ì‹œì‘",
                                            command=self.start_analysis,
                                            font=("Consolas", 24, "bold"),
                                            fg_color="#2196F3",
                                            hover_color="#1976D2",
                                            height=50,
                                            state="disabled")
            self.analyze_btn.pack(side="left", padx=(10, 5), fill="x", expand=True)

            # Result í´ë” ì—´ê¸° ë²„íŠ¼
            self.open_result_btn = ctk.CTkButton(button_frame, text="ğŸ“ Result í´ë”",
                                                command=self.open_result_folder,
                                                font=("Consolas", 21, "bold"),
                                                fg_color="#FF6B00",
                                                hover_color="#E65100",
                                                text_color="white",
                                                height=50)
            self.open_result_btn.pack(side="left", padx=(5, 10), fill="x", expand=True)

            # ì§„í–‰ ìƒíƒœ í‘œì‹œ
            self.status_label = ctk.CTkLabel(self.root, text="ë°ì´í„° íŒŒì¼ì„ ì„ íƒí•´ì£¼ì„¸ìš”.",
                                            font=("Consolas", 20))
            self.status_label.pack(pady=6)

            # ì €ì¥ ê²½ë¡œ í‘œì‹œ ë ˆì´ë¸”
            self.saved_path_label = ctk.CTkLabel(self.root, text="",
                                                font=("Consolas", 17),
                                                text_color="#1E88E5",
                                                wraplength=680)
            self.saved_path_label.pack(pady=0)

            # ì§„í–‰ í‘œì‹œê¸°
            self.progress = ctk.CTkProgressBar(self.root, mode='indeterminate',
                                              width=550, height=28)
            self.progress.set(0)
            
        def browse_file(self):
            """íŒŒì¼ ì„ íƒ ëŒ€í™”ìƒì"""
            filename = filedialog.askopenfilename(
                title="ë°ì´í„° íŒŒì¼ ì„ íƒ",
                filetypes=[("Data files", "*.xlsx *.xls *.csv"), ("Excel files", "*.xlsx *.xls"), ("CSV files", "*.csv"), ("All files", "*.*")]
            )
            
            if filename:
                self.selected_file = filename
                self.file_entry.delete(0, "end")
                self.file_entry.insert(0, filename)
                self.analyze_btn.configure(state="normal")
                self.status_label.configure(text="ë¶„ì„ ì¤€ë¹„ ì™„ë£Œ. 'ë¶„ì„ ì‹œì‘' ë²„íŠ¼ì„ í´ë¦­í•˜ì„¸ìš”.",
                                           text_color="#4CAF50")
        
        def start_analysis(self):
            """ë¶„ì„ ì‹œì‘"""
            if not self.selected_file:
                messagebox.showerror("ì˜¤ë¥˜", "íŒŒì¼ì„ ì„ íƒí•´ì£¼ì„¸ìš”.")
                return

            # ë²„íŠ¼ ë¹„í™œì„±í™”
            self.analyze_btn.configure(state="disabled")
            self.status_label.configure(text="ë¶„ì„ ì¤‘...", text_color="#2196F3")
            self.progress.pack(pady=10)
            self.progress.start()

            # ë³„ë„ ìŠ¤ë ˆë“œì—ì„œ ë¶„ì„ ì‹¤í–‰
            thread = threading.Thread(target=self.run_analysis)
            thread.start()
        
        def run_analysis(self):
            """ë¶„ì„ ì‹¤í–‰ (ë³„ë„ ìŠ¤ë ˆë“œ)"""
            try:
                output_path = process_file(self.selected_file)
                self.output_file = output_path

                # ì„±ê³µ ë©”ì‹œì§€
                self.root.after(0, self.analysis_complete)
            except Exception as e:
                # ì˜¤ë¥˜ ë©”ì‹œì§€
                self.root.after(0, lambda: self.analysis_error(str(e)))
        
        def analysis_complete(self):
            """ë¶„ì„ ì™„ë£Œ"""
            self.progress.stop()
            self.progress.pack_forget()
            self.status_label.configure(text="ë¶„ì„ ì™„ë£Œ! ê²°ê³¼ê°€ Result í´ë”ì— ì €ì¥ë˜ì—ˆìŠµë‹ˆë‹¤.",
                                       text_color="#4CAF50")

            # ì €ì¥ëœ íŒŒì¼ ê²½ë¡œ í‘œì‹œ
            if self.output_file:
                abs_path = os.path.abspath(self.output_file)
                self.saved_path_label.configure(text=f"ì €ì¥ ê²½ë¡œ: {abs_path}")

            self.analyze_btn.configure(state="normal")

            messagebox.showinfo("ì„±ê³µ",
                              f"ë¶„ì„ì´ ì„±ê³µì ìœ¼ë¡œ ì™„ë£Œë˜ì—ˆìŠµë‹ˆë‹¤!\n\n"
                              f"ê²°ê³¼ íŒŒì¼: {self.output_file}")
        
        def analysis_error(self, error_msg):
            """ë¶„ì„ ì˜¤ë¥˜"""
            self.progress.stop()
            self.progress.pack_forget()
            self.status_label.configure(text="ë¶„ì„ ì‹¤íŒ¨", text_color="#F44336")
            self.analyze_btn.configure(state="normal")

            messagebox.showerror("ì˜¤ë¥˜", f"ë¶„ì„ ì¤‘ ì˜¤ë¥˜ê°€ ë°œìƒí–ˆìŠµë‹ˆë‹¤:\n\n{error_msg}")

        def open_result_folder(self):
            """Result í´ë”ë¥¼ íƒìƒ‰ê¸°ë¡œ ì—´ê¸°"""
            result_dir = 'Result'
            if not os.path.exists(result_dir):
                os.makedirs(result_dir)

            # Windowsì—ì„œ í´ë” ì—´ê¸°
            abs_path = os.path.abspath(result_dir)
            os.startfile(abs_path)
    
    # GUI ì‹¤í–‰
    root = ctk.CTk()
    app = HysteresisApp(root)
    root.mainloop()

if __name__ == '__main__':
    # ì»¤ë§¨ë“œë¼ì¸ ì¸ìˆ˜ê°€ ìˆìœ¼ë©´ CLI ëª¨ë“œ, ì—†ìœ¼ë©´ GUI ëª¨ë“œ
    if len(sys.argv) > 1:
        main_cli()
    else:
        main_gui()
