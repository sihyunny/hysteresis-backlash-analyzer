#!/usr/bin/env python3.11
# -*- coding: utf-8 -*-

"""
히스테리시스 곡선 분석 프로그램 - 최종 버전
다양한 데이터 형식에 대응 (큰 변화 + 연속적 변화)
토크 강성 계산 추가
"""

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import sys
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment

def find_hysteresis_loop_v2(x_data, y_data):
    """
    개선된 히스테리시스 루프 감지
    방법 1: X값 최대 지점 기준 (test001.xlsx 타입) - 우선 사용
    방법 2: 큰 X값 변화 감지 (test1.xlsx 타입) - 방법 1이 부적합할 때 사용
    """
    # 방법 1: X값 최대 지점 기준 (기본 방법)
    x_max_idx = np.argmax(x_data)
    x_max = x_data[x_max_idx]
    
    # X값이 최대 지점 근처에 여러 개 있을 수 있으므로 첫 번째 최대값 찾기
    first_max_idx = np.where(x_data == x_max)[0][0]
    
    ascending_start = 0
    descending_start = first_max_idx
    descending_end = len(x_data)
    
    # 방법 1이 적합한지 확인 (상승 경로와 하강 경로가 모두 충분한지)
    if descending_start > 10 and (descending_end - descending_start) > 10:
        return ascending_start, descending_start, descending_end, "X 최대값 기준"
    
    # 방법 2: 큰 변화 감지 (방법 1이 부적합할 때)
    x_diff = np.diff(x_data)
    
    large_increases = []
    large_decreases = []
    
    for i in range(len(x_diff)):
        if x_diff[i] > 1.0:
            large_increases.append(i)
        if x_diff[i] < -1.0:
            large_decreases.append(i)
    
    # 큰 변화가 있으면 방법 2 사용
    if len(large_increases) > 0 and len(large_decreases) > 0:
        ascending_start = large_increases[0] + 1
        descending_start = None
        
        for dec_idx in large_decreases:
            if dec_idx > ascending_start:
                descending_start = dec_idx + 1
                break
        
        if descending_start is not None:
            descending_end = len(x_data)
            for inc_idx in large_increases:
                if inc_idx > descending_start:
                    descending_end = inc_idx + 1
                    break
            
            return ascending_start, descending_start, descending_end, "큰 변화 감지"
    
    # 둘 다 실패하면 방법 1의 결과 반환
    return 0, first_max_idx, len(x_data), "X 최대값 기준 (기본)"

def calculate_backlash(file_path):
    """히스테리시스 루프를 자동 감지하여 백래쉬 계산"""
    # 파일 확장자에 따라 적절한 함수로 읽기
    file_ext = os.path.splitext(file_path)[1].lower()
    if file_ext == '.csv':
        df = pd.read_csv(file_path, header=None)
    else:  # .xlsx, .xls 등
        df = pd.read_excel(file_path, header=None)

    # A열(0번 열)이 0인 행들을 제거
    df = df[df[0] != 0]

    x_data = df[0].values
    y_data = df[1].values

    # X축 기준값 계산
    x_max = x_data.max()
    x_min = x_data.min()
    
    # X축 3% 지점과 -3% 지점 계산
    x_3_percent = x_max * 0.03
    x_minus_3_percent = x_min * 0.03
    
    # 히스테리시스 루프 자동 감지
    ascending_start, descending_start, descending_end, method = find_hysteresis_loop_v2(x_data, y_data)
    
    if ascending_start is None or descending_start is None:
        return None
    
    # 상승 경로
    ascending_x = x_data[ascending_start:descending_start]
    ascending_y = y_data[ascending_start:descending_start]
    
    # 하강 경로
    descending_x = x_data[descending_start:descending_end]
    descending_y = y_data[descending_start:descending_end]
    
    # 선형 보간을 사용하여 특정 X값에서의 Y값 찾기
    def find_y_at_x(x_arr, y_arr, target_x):
        if len(x_arr) == 0:
            return None
        
        # target_x에 정확히 일치하는 점이 있는지 확인
        exact_match = np.where(np.abs(x_arr - target_x) < 0.01)[0]
        if len(exact_match) > 0:
            return y_arr[exact_match[0]]
        
        # target_x보다 작거나 같은 점과 크거나 같은 점 찾기
        lower_indices = np.where(x_arr <= target_x)[0]
        upper_indices = np.where(x_arr >= target_x)[0]
        
        if len(lower_indices) > 0 and len(upper_indices) > 0:
            lower_idx = lower_indices[np.argmax(x_arr[lower_indices])]
            upper_idx = upper_indices[np.argmin(x_arr[upper_indices])]
            
            x1, y1 = x_arr[lower_idx], y_arr[lower_idx]
            x2, y2 = x_arr[upper_idx], y_arr[upper_idx]
            
            if x2 != x1:
                y_interpolated = y1 + (y2 - y1) * (target_x - x1) / (x2 - x1)
                return y_interpolated
            else:
                return y1
        elif len(lower_indices) > 0:
            return y_arr[lower_indices[-1]]
        elif len(upper_indices) > 0:
            return y_arr[upper_indices[0]]
        
        return None
    
    # back1: 전체 데이터에서 X축 3% 지점의 Y값 (선형 보간 사용)
    # X=3%를 지나가는 모든 구간에서 선형 보간한 후, 가장 큰 Y값 선택
    back1 = None
    back1_candidates = []

    for i in range(len(x_data) - 1):
        x1, x2 = x_data[i], x_data[i + 1]
        y1, y2 = y_data[i], y_data[i + 1]

        # X=3% 지점이 이 구간 [x1, x2] 사이에 있는지 확인
        if (x1 <= x_3_percent <= x2) or (x2 <= x_3_percent <= x1):
            # 선형 보간으로 Y값 계산
            if x2 != x1:
                y_interpolated = y1 + (y2 - y1) * (x_3_percent - x1) / (x2 - x1)
                back1_candidates.append(y_interpolated)

    # 모든 후보 중 가장 큰 Y값 선택
    if len(back1_candidates) > 0:
        back1 = max(back1_candidates)
    else:
        # 후보가 없으면 기존 방식 사용
        back1 = find_y_at_x(x_data, y_data, x_3_percent)

    # back2: 전체 데이터에서 X축 -3% 지점의 Y값 (선형 보간 사용)
    # X=-3%를 지나가는 모든 구간에서 선형 보간한 후, 가장 작은 Y값 선택
    back2 = None
    back2_candidates = []

    for i in range(len(x_data) - 1):
        x1, x2 = x_data[i], x_data[i + 1]
        y1, y2 = y_data[i], y_data[i + 1]

        # X=-3% 지점이 이 구간 [x1, x2] 사이에 있는지 확인
        if (x1 <= x_minus_3_percent <= x2) or (x2 <= x_minus_3_percent <= x1):
            # 선형 보간으로 Y값 계산
            if x2 != x1:
                y_interpolated = y1 + (y2 - y1) * (x_minus_3_percent - x1) / (x2 - x1)
                back2_candidates.append(y_interpolated)

    # 모든 후보 중 가장 작은 Y값 선택
    if len(back2_candidates) > 0:
        back2 = min(back2_candidates)
    else:
        # 후보가 없으면 기존 방식 사용
        back2 = find_y_at_x(x_data, y_data, x_minus_3_percent)

    # 토크 강성 계산을 위한 tolerance 변수 (기존 위치 유지)
    tolerance = abs(x_max - x_min) * 0.01  # 1% 허용 오차
    
    # 백래쉬 계산
    backlash = None
    if back1 is not None and back2 is not None:
        backlash = abs(back1 - back2)
    
    # 토크 강성 계산
    # X=50% 지점과 X=100% 지점 계산
    x_50_percent = x_max * 0.5
    x_100_percent = x_max
    
    # X=100% 지점의 Y값 (가장 가까운 실제 데이터 포인트)
    y_100_percent = find_y_at_x(x_data, y_data, x_100_percent)
    
    # X=50% 지점의 두 Y값 찾기 (상승 경로와 하강 경로)
    mask_50 = np.abs(x_data - x_50_percent) <= tolerance
    y_50_percent_min = None
    y_50_percent_all = []
    
    if np.any(mask_50):
        y_50_percent_all = y_data[mask_50].tolist()
        y_50_percent_min = np.min(y_data[mask_50])  # 더 작은 Y값
    else:
        # 허용 오차 내에 데이터가 없으면 선형 보간 사용
        y_50_percent_min = find_y_at_x(x_data, y_data, x_50_percent)
        if y_50_percent_min is not None:
            y_50_percent_all = [y_50_percent_min]
    
    # 우측 토크 강성 계산: b/a
    b = x_100_percent - x_50_percent  # X값의 차이
    a = None
    torque_stiffness = None
    
    if y_100_percent is not None and y_50_percent_min is not None:
        a = y_100_percent - y_50_percent_min  # Y값의 차이
        if a != 0:
            torque_stiffness = b / a
    
    # 좌측 토크 강성 계산 (X=-50%, X=-100%)
    x_minus_50_percent = x_min * 0.5
    x_minus_100_percent = x_min
    
    # X=-100% 지점의 Y값
    y_minus_100_percent = find_y_at_x(x_data, y_data, x_minus_100_percent)
    
    # X=-50% 지점의 두 Y값 찾기 (상승 경로와 하강 경로)
    mask_minus_50 = np.abs(x_data - x_minus_50_percent) <= tolerance
    y_minus_50_percent_max = None
    y_minus_50_percent_all = []
    
    if np.any(mask_minus_50):
        y_minus_50_percent_all = y_data[mask_minus_50].tolist()
        y_minus_50_percent_max = np.max(y_data[mask_minus_50])  # 더 큰 Y값
    else:
        # 허용 오차 내에 데이터가 없으면 선형 보간 사용
        y_minus_50_percent_max = find_y_at_x(x_data, y_data, x_minus_50_percent)
        if y_minus_50_percent_max is not None:
            y_minus_50_percent_all = [y_minus_50_percent_max]
    
    # 좌측 토크 강성 계산: b_left/a_left
    b_left = x_minus_50_percent - x_minus_100_percent  # X값의 차이 (절대값)
    a_left = None
    torque_stiffness_left = None
    
    if y_minus_100_percent is not None and y_minus_50_percent_max is not None:
        a_left = y_minus_100_percent - y_minus_50_percent_max  # Y값의 차이
        if a_left != 0:
            torque_stiffness_left = abs(b_left / a_left)
    
    return {
        'x_data': x_data,
        'y_data': y_data,
        'ascending_x': ascending_x,
        'ascending_y': ascending_y,
        'descending_x': descending_x,
        'descending_y': descending_y,
        'x_max': x_max,
        'x_min': x_min,
        'x_3_percent': x_3_percent,
        'x_minus_3_percent': x_minus_3_percent,
        'back1': back1,
        'back2': back2,
        'backlash': backlash,
        'method': method,
        'x_50_percent': x_50_percent,
        'x_100_percent': x_100_percent,
        'y_50_percent_min': y_50_percent_min,
        'y_50_percent_all': y_50_percent_all,
        'y_100_percent': y_100_percent,
        'b': b,
        'a': a,
        'torque_stiffness': torque_stiffness,
        'x_minus_50_percent': x_minus_50_percent,
        'x_minus_100_percent': x_minus_100_percent,
        'y_minus_50_percent_max': y_minus_50_percent_max,
        'y_minus_50_percent_all': y_minus_50_percent_all,
        'y_minus_100_percent': y_minus_100_percent,
        'b_left': b_left,
        'a_left': a_left,
        'torque_stiffness_left': torque_stiffness_left
    }

def plot_and_save(result, output_path, input_filename=None):
    """그래프 그리기 및 저장"""
    if not result:
        print("분석 결과가 없습니다.")
        return False
    
    ascending_x = result['ascending_x']
    ascending_y = result['ascending_y']
    descending_x = result['descending_x']
    descending_y = result['descending_y']
    x_3_percent = result['x_3_percent']
    x_minus_3_percent = result['x_minus_3_percent']
    back1 = result['back1']
    back2 = result['back2']
    backlash = result['backlash']
    
    # 우측 토크 강성 관련 변수
    x_50_percent = result.get('x_50_percent')
    x_100_percent = result.get('x_100_percent')
    y_50_percent_min = result.get('y_50_percent_min')
    y_100_percent = result.get('y_100_percent')
    b = result.get('b')
    a = result.get('a')
    torque_stiffness = result.get('torque_stiffness')
    
    # 좌측 토크 강성 관련 변수
    x_minus_50_percent = result.get('x_minus_50_percent')
    x_minus_100_percent = result.get('x_minus_100_percent')
    y_minus_50_percent_max = result.get('y_minus_50_percent_max')
    y_minus_100_percent = result.get('y_minus_100_percent')
    b_left = result.get('b_left')
    a_left = result.get('a_left')
    torque_stiffness_left = result.get('torque_stiffness_left')
    
    # 그래프 생성
    graph_left = 0.08  # 그래프 왼쪽 여백
    fig = plt.figure(figsize=(12, 8.5))

    # 전체 히스테리시스 곡선을 파란색으로 표시
    x_data = result['x_data']
    y_data = result['y_data']
    plt.plot(x_data, y_data, 'b-', linewidth=1.2, alpha=0.8, label='Hysteresis Curve')
    plt.scatter(x_data, y_data, c='orange', s=10, alpha=1.0, zorder=3, label='Data Points')
    
    # X축 3% 지점과 -3% 지점 표시
    if back1 is not None:
        plt.scatter([x_3_percent], [back1], c='red', s=100, marker='o',
                    edgecolors='black', linewidths=2,
                    label=f'X=3%: Y={back1:.4f}', zorder=5)
        plt.axvline(x=x_3_percent, color='red', linestyle='--', linewidth=2.0, alpha=1.0)
        # 3% 텍스트 표시
        plt.text(x_3_percent, result['y_data'].min() * 0.95, '3%',
                fontsize=10, fontweight='bold', color='red',
                horizontalalignment='center', verticalalignment='top')

    if back2 is not None:
        plt.scatter([x_minus_3_percent], [back2], c='green', s=100, marker='s',
                    edgecolors='black', linewidths=2,
                    label=f'X=-3%: Y={back2:.4f}', zorder=5)
        plt.axvline(x=x_minus_3_percent, color='green', linestyle='--', linewidth=1.5, alpha=0.6)
        # -3% 텍스트 표시
        plt.text(x_minus_3_percent, result['y_data'].min() * 0.95, '-3%',
                fontsize=10, fontweight='bold', color='green',
                horizontalalignment='center', verticalalignment='top')
    
    # 백래쉬를 시각적으로 표시 (수평선 + 양방향 화살표)
    if back1 is not None and back2 is not None:
        # back1에서 수평선 (오른쪽으로)
        x_line_end = result['x_data'].max() * 0.2  # 그래프 중간까지
        plt.plot([x_3_percent, x_line_end], [back1, back1],
                'red', linewidth=2.0, linestyle='-', alpha=1.0, zorder=4)
        
        # back2에서 수평선 (오른쪽으로)
        plt.plot([x_minus_3_percent, x_line_end], [back2, back2], 
                'green', linewidth=1.5, linestyle='-', alpha=0.7, zorder=4)
        
        # 양방향 화살표 (수평선 끝에서)
        arrow_x = x_line_end * 0.9  # 화살표 위치
        plt.annotate('', xy=(arrow_x, back1), xytext=(arrow_x, back2),
                    arrowprops=dict(arrowstyle='<->', color='gray', lw=2, 
                                  mutation_scale=25, shrinkA=0, shrinkB=0),
                    zorder=6)
        
        # 백래쉬 텍스트 (화살표 옆)
        text_x = arrow_x * 0.6
        text_y = (back1 + back2) / 2
        plt.text(text_x, text_y, f'Back Lash\n={backlash:.2f} arc Min',
                fontsize=14, fontweight='bold',
                verticalalignment='center',
                bbox=dict(boxstyle='round', facecolor='yellow', alpha=0.9, 
                          edgecolor='black', linewidth=2.5),
                zorder=7)
    
    # 토크 강성 시각화
    if x_50_percent is not None and x_100_percent is not None:
        # X=50% 지점에 파란색 수직 점선 (Y=0 이상에서만)
        y_max = y_50_percent_min
        plt.plot([x_50_percent, x_50_percent], [0, y_max], color='blue', linestyle='--', linewidth=1.5, alpha=0.6, zorder=2)
        # 50% 텍스트 표시
        plt.text(x_50_percent, -0.2, '50%', 
                fontsize=10, fontweight='bold', color='blue',
                horizontalalignment='center', verticalalignment='top')
        
        # X=100% 지점에 파란색 수직 점선 (Y=0 이상에서만)
        plt.plot([x_100_percent, x_100_percent], [0, y_100_percent], color='blue', linestyle='--', linewidth=1.5, alpha=0.6,  zorder=2)
        # 100% 텍스트 표시
        plt.text(x_100_percent, -0.2, '100%', 
                fontsize=10, fontweight='bold', color='blue',
                horizontalalignment='center', verticalalignment='top')
        
        if y_50_percent_min is not None and y_100_percent is not None:

            # b 값 표시 (X축 방향, 하늘색)
            mid_y = y_50_percent_min
            plt.annotate('', xy=(x_100_percent, mid_y), xytext=(x_50_percent, mid_y),
                        arrowprops=dict(arrowstyle='<->', color='black', lw=1.5, 
                                      mutation_scale=20, shrinkA=0, shrinkB=0),
                        zorder=6)
            plt.text((x_50_percent + x_100_percent) / 2, mid_y * 1.02, f'b={b:.2f}',
                    fontsize=10, fontweight='bold', color='black',
                    horizontalalignment='center')
            
            # a 값 표시 (Y축 방향, 주황색)
            mid_x = x_100_percent
            plt.annotate('', xy=(mid_x, y_100_percent), xytext=(mid_x, y_50_percent_min),
                        arrowprops=dict(arrowstyle='<->', color='black', lw=1.5, 
                                      mutation_scale=20, shrinkA=0, shrinkB=0),
                        zorder=6)
            plt.text(mid_x * 1.02, (y_50_percent_min + y_100_percent) / 2, f'a={a:.2f}',
                    fontsize=10, fontweight='bold', color='black',
                    verticalalignment='center')
            
            # 우측 토크 강성 텍스트 표시
            if torque_stiffness is not None:
                plt.text(x_100_percent*0.98, mid_y * 0.9, 
                        f'Torsional Stiffness = b/a\n= {torque_stiffness:.4f}',
                        #transform=plt.gca().transAxes,
                        fontsize=12, fontweight='bold',
                        horizontalalignment='right',
                        verticalalignment='top', color='blue')
    
    # 좌측 토크 강성 시각화
    if x_minus_50_percent is not None and x_minus_100_percent is not None:
        # X=-50% 지점에 파란색 수직 점선 (Y=0 이하에서만)
        y_min = result['y_data'].min()
        plt.plot([x_minus_50_percent, x_minus_50_percent], [y_minus_50_percent_max, 0], color='blue', linestyle='--', linewidth=1.5, alpha=0.6, zorder=2)
        # -50% 텍스트 표시
        plt.text(x_minus_50_percent, -0.2, '50%', 
                fontsize=10, fontweight='bold', color='blue',
                horizontalalignment='center', verticalalignment='top')
        
        # X=-100% 지점에 파란색 수직 점선 (Y=0 이하에서만)
        plt.plot([x_minus_100_percent, x_minus_100_percent], [y_min, 0], color='blue', linestyle='--', linewidth=1.5, alpha=0.6,  zorder=2)
        # -100% 텍스트 표시
        plt.text(x_minus_100_percent, -0.2, '100%', 
                fontsize=10, fontweight='bold', color='blue',
                horizontalalignment='center', verticalalignment='top')
        
        if y_minus_50_percent_max is not None and y_minus_100_percent is not None:
            # b_left 값 표시 (X축 방향)
            mid_y_left = y_minus_50_percent_max
            plt.annotate('', xy=(x_minus_50_percent, mid_y_left), xytext=(x_minus_100_percent, mid_y_left),
                        arrowprops=dict(arrowstyle='<->', color='black', lw=1.5, 
                                      mutation_scale=20, shrinkA=0, shrinkB=0),
                        zorder=6)
            plt.text((x_minus_50_percent + x_minus_100_percent) / 2, mid_y_left*1.06 , f'b={abs(b_left):.2f}',
                    fontsize=10, fontweight='bold', color='black',
                    horizontalalignment='center')
            
            # a_left 값 표시 (Y축 방향)
            mid_x_left = x_minus_100_percent
            plt.annotate('', xy=(mid_x_left, y_minus_100_percent), xytext=(mid_x_left, y_minus_50_percent_max),
                        arrowprops=dict(arrowstyle='<->', color='black', lw=1.5, 
                                      mutation_scale=20, shrinkA=0, shrinkB=0),
                        zorder=6)
            plt.text(mid_x_left * 1, (y_minus_50_percent_max + y_minus_100_percent) / 2, f'a={abs(a_left):.2f}',
                    fontsize=10, fontweight='bold', color='black',
                    verticalalignment='center', horizontalalignment='right')
            
            # 좌측 토크 강성 텍스트 표시
            if torque_stiffness_left is not None:
                plt.text(mid_x_left*0.98,y_minus_50_percent_max*0.9, f'Torsional Stiffness = b/a\n= {torque_stiffness_left:.4f}',
                        fontsize=12, fontweight='bold',
                        horizontalalignment='left',
                        verticalalignment='bottom',color='blue')
    
    plt.xlabel('Torque (Nm)', fontsize=14, fontweight='bold')
    plt.ylabel('Angle (arc min)', fontsize=14, fontweight='bold')
    plt.title('Hysteresis Curve', fontsize=16, fontweight='bold')
    
    # 격자선 설정
    from matplotlib.ticker import MultipleLocator
    ax = plt.gca()
    
    # X축 격자선: 주 10 단위, 보조 2.5 단위
    ax.xaxis.set_major_locator(MultipleLocator(10))
    ax.xaxis.set_minor_locator(MultipleLocator(2.5))
    
    # Y축 격자선: 주 10 단위, 보조 2.5 단위
    ax.yaxis.set_major_locator(MultipleLocator(5))
    ax.yaxis.set_minor_locator(MultipleLocator(2.5))
    
    # 주 격자선 (Major grid) - 회색, 연하고 얘게
    plt.grid(True, which='major', alpha=0.5, linestyle='-', linewidth=0.6, color='gray')
    
    # 보조 격자선 (Minor grid) - 연한 회색, 더 연하고 얘게
    plt.grid(True, which='minor', alpha=0.25, linestyle='-', linewidth=0.3, color='lightgray')
    
    plt.legend(fontsize=11, loc='best')
    plt.axhline(y=0, color='k', linestyle='-', linewidth=2, alpha=1)
    plt.axvline(x=0, color='k', linestyle='-', linewidth=2, alpha=1)

    # Y축을 0 중심으로 대칭 설정
    y_max_abs = max(abs(result['y_data'].min()), abs(result['y_data'].max()))
    plt.ylim(-y_max_abs * 1.05, y_max_abs * 1.05)  # 5% 여유 공간

    # 엑셀 파일명을 그래프 왼쪽 끝과 정렬하여 표시 (초록색 글씨)
    if input_filename is not None:
        fig.text(graph_left, 0.99, f'File: {input_filename}',
                 fontsize=10, fontweight='bold',
                 color='green',
                 verticalalignment='top',
                 horizontalalignment='left')

    # 생성 날짜/시간을 파일명 바로 밑에 좌측 정렬하여 표시
    creation_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    fig.text(graph_left, 0.965, creation_datetime,
             fontsize=11, fontweight='bold',
             fontfamily='monospace',
             color="#165AF7",  # 진한 파란색 (darkblue)
             verticalalignment='top',
             horizontalalignment='left')

    # tight_layout은 GridSpec 사용 시 제거 (이미 레이아웃 설정됨)
    plt.savefig(output_path, dpi=300, bbox_inches='tight')
    plt.close()
    
    return True

def log_backlash(input_filename, backlash_value):
    """백래쉬 값을 엑셀 로그 파일에 기록 (서식 적용)"""
    result_dir = 'Result'
    log_file = os.path.join(result_dir, 'backlash_log.xlsx')

    # 현재 날짜/시간
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # 엑셀 파일이 없으면 새로 생성
    if not os.path.exists(log_file):
        wb = Workbook()
        ws = wb.active
        ws.title = "Backlash Log"

        # 헤더 작성
        headers = ["DateTime", "FileName", "Backlash(arc min)"]
        ws.append(headers)

        # 헤더 서식 적용 (1행)
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = Font(bold=True, color="0000FF")  # 볼드, 파란색
            cell.alignment = Alignment(horizontal="center", vertical="center")

        wb.save(log_file)

    # 기존 파일에 데이터 추가
    wb = load_workbook(log_file)
    ws = wb.active

    # 새 행 추가
    new_row = [timestamp, input_filename, f"{backlash_value:.4f}"]
    ws.append(new_row)

    # 새로 추가된 행의 셀에 중앙 정렬 적용
    row_num = ws.max_row
    for col_num in range(1, 4):
        cell = ws.cell(row=row_num, column=col_num)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    wb.save(log_file)
    print(f"백래쉬 로그 기록 완료: {log_file}")

def process_file(input_file):
    """파일 분석 처리 함수"""
    
    if not os.path.exists(input_file):
        print(f"오류: 파일을 찾을 수 없습니다: {input_file}")
        sys.exit(1)
    
    # 엑셀 파일명 추출
    input_filename = os.path.basename(input_file)
    
    # Result 폴더 생성
    result_dir = 'Result'
    if not os.path.exists(result_dir):
        os.makedirs(result_dir)
        print(f"Result 폴더를 생성했습니다: {result_dir}")
    
    # 출력 파일 경로 설정
    if len(sys.argv) >= 3:
        output_file = os.path.join(result_dir, os.path.basename(sys.argv[2]))
    else:
        base_name = os.path.splitext(input_filename)[0]
        datetime_str = datetime.now().strftime("%Y.%m.%d_%Hh%Mm%Ss")
        output_file = os.path.join(result_dir, f"{datetime_str}-{base_name}.png")
    
    print(f"입력 파일: {input_file}")
    print(f"출력 파일: {output_file}")
    print("\n분석 중...")
    
    try:
        # 분석 수행
        result = calculate_backlash(input_file)
        
        if result:
            # 결과 출력
            print("\n========== 히스테리시스 곡선 분석 결과 ==========")
            print(f"감지 방법: {result['method']}")
            print(f"데이터 포인트 수: {len(result['x_data'])}")
            print(f"X축 범위: {result['x_data'].min():.2f} ~ {result['x_data'].max():.2f}")
            print(f"Y축 범위: {result['y_data'].min():.2f} ~ {result['y_data'].max():.2f}")
            print(f"\n상승 경로: {len(result['ascending_x'])} 포인트")
            print(f"하강 경로: {len(result['descending_x'])} 포인트")
            print(f"\nX축 최대값 (100%): {result['x_max']:.4f}")
            print(f"X축 최소값 (-100%): {result['x_min']:.4f}")
            print(f"\nX축 3% 지점: {result['x_3_percent']:.4f}")
            print(f"X축 -3% 지점: {result['x_minus_3_percent']:.4f}")
            print(f"\nX=3% 지점의 Y값: {result['back1']:.4f}")
            print(f"X=-3% 지점의 Y값: {result['back2']:.4f}")
            print(f"\n백래쉬 (Backlash): {result['backlash']:.4f}")
            
            # 우측 토크 강성 결과 출력
            if result.get('torque_stiffness') is not None:
                print(f"\n========== 우측 토크 강성 계산 결과 ==========")
                print(f"X=50% 지점: {result['x_50_percent']:.4f}")
                print(f"X=100% 지점: {result['x_100_percent']:.4f}")
                print(f"Y (X=50%, 최소값): {result['y_50_percent_min']:.4f}")
                print(f"Y (X=100%): {result['y_100_percent']:.4f}")
                print(f"b (X값 차이): {result['b']:.4f}")
                print(f"a (Y값 차이): {result['a']:.4f}")
                print(f"토크 강성 (b/a): {result['torque_stiffness']:.4f}")
            
            # 좌측 토크 강성 결과 출력
            if result.get('torque_stiffness_left') is not None:
                print(f"\n========== 좌측 토크 강성 계산 결과 ==========")
                print(f"X=-50% 지점: {result['x_minus_50_percent']:.4f}")
                print(f"X=-100% 지점: {result['x_minus_100_percent']:.4f}")
                print(f"Y (X=-50%, 최대값): {result['y_minus_50_percent_max']:.4f}")
                print(f"Y (X=-100%): {result['y_minus_100_percent']:.4f}")
                print(f"b_left (X값 차이): {abs(result['b_left']):.4f}")
                print(f"a_left (Y값 차이): {abs(result['a_left']):.4f}")
                print(f"토크 강성 (b/a): {result['torque_stiffness_left']:.4f}")
            
            print("=" * 50)
            
            # 그래프 저장
            if plot_and_save(result, output_file, input_filename):
                print(f"\n그래프가 성공적으로 저장되었습니다: {output_file}")

                # 백래쉬 값 로그 기록
                log_backlash(input_filename, result['backlash'])

                return output_file  # 출력 파일 경로 반환
            else:
                print("\n그래프 저장에 실패했습니다.")
                sys.exit(1)
        else:
            print("분석에 실패했습니다.")
            sys.exit(1)

    except Exception as e:
        print(f"\n오류 발생: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

def main_cli():
    """커맨드라인 모드"""
    if len(sys.argv) < 2:
        print("사용법: python3.11 graph_output.py <데이터파일경로(.xlsx/.xls/.csv)> [출력파일경로]")
        sys.exit(1)
    
    input_file = sys.argv[1]
    process_file(input_file)

def main_gui():
    """간단한 GUI 모드 - CustomTkinter"""
    import customtkinter as ctk
    from tkinter import filedialog, messagebox
    import threading

    # CustomTkinter 테마 설정
    ctk.set_appearance_mode("dark")
    ctk.set_default_color_theme("blue")
    
    class HysteresisApp:
        def __init__(self, root):
            self.root = root
            self.root.title("Hysteresis Curve\nBacklash Analyzer")
            self.root.geometry("750x650")
            self.root.resizable(False, False)

            # 선택된 파일 경로
            self.selected_file = None
            # 저장된 출력 파일 경로
            self.output_file = None

            # UI 구성
            self.create_widgets()
            
        def create_widgets(self):
            # 제목 및 로고 프레임
            header_frame = ctk.CTkFrame(self.root, fg_color="transparent")
            header_frame.pack(pady=15, padx=20, fill="x")

            # 제목 (좌측) - 2줄
            title_label = ctk.CTkLabel(header_frame, text="Hysteresis Curve\nBacklash Analyzer",
                                       font=("Consolas", 28, "bold"),
                                       justify="left")
            title_label.pack(side="left")

            # 사용 가이드 프레임
            guide_frame = ctk.CTkFrame(self.root)
            guide_frame.pack(pady=10, padx=0, fill="x")

            guide_title = ctk.CTkLabel(guide_frame, text="📢 GUIDE",
                                       font=("Consolas", 19))
            guide_title.pack(anchor="w", padx=15, pady=(10, 5))

            guide_text = (
                " ৹ 입력: Excel/CSV 파일 (A열=토크, B열=각도)\n\n"
                " ৹ 분석: 히스테리시스 곡선 생성 + 백래쉬 + 토크 강성 자동 계산\n\n"
                " ৹ 출력: Result 폴더에 그래프 + 로그 저장"
            )

            guide_label = ctk.CTkLabel(guide_frame, text=guide_text,
                                       font=("돋움", 16),
                                       justify="left")
            guide_label.pack(anchor="w", padx=15, pady=(0, 10))

            # 파일 선택 프레임
            file_frame = ctk.CTkFrame(self.root)
            file_frame.pack(pady=12, padx=20, fill="x")

            file_label = ctk.CTkLabel(file_frame, text="데이터 파일:",
                                     font=("Consolas", 20, "bold"))
            file_label.pack(side="left", padx=(10, 5))

            browse_btn = ctk.CTkButton(file_frame, text="파일 선택",
                                      command=self.browse_file,
                                      font=("Consolas", 18, "bold"),
                                      fg_color="#4CAF50",
                                      hover_color="#45a049",
                                      width=130, height=50)
            browse_btn.pack(side="right", padx=(5, 10))

            self.file_entry = ctk.CTkEntry(file_frame, font=("Consolas", 18),
                                          height=50)
            self.file_entry.pack(side="left", padx=5, fill="x", expand=True)

            # 버튼 프레임 (분석 시작 + Result 폴더 열기)
            button_frame = ctk.CTkFrame(self.root, fg_color="transparent")
            button_frame.pack(pady=12, padx=20, fill="x")

            # 분석 버튼
            self.analyze_btn = ctk.CTkButton(button_frame, text="분석 시작",
                                            command=self.start_analysis,
                                            font=("Consolas", 24, "bold"),
                                            fg_color="#2196F3",
                                            hover_color="#1976D2",
                                            height=50,
                                            state="disabled")
            self.analyze_btn.pack(side="left", padx=(10, 5), fill="x", expand=True)

            # Result 폴더 열기 버튼
            self.open_result_btn = ctk.CTkButton(button_frame, text="📁 Result 폴더",
                                                command=self.open_result_folder,
                                                font=("Consolas", 21, "bold"),
                                                fg_color="#FF6B00",
                                                hover_color="#E65100",
                                                text_color="white",
                                                height=50)
            self.open_result_btn.pack(side="left", padx=(5, 10), fill="x", expand=True)

            # 진행 상태 표시
            self.status_label = ctk.CTkLabel(self.root, text="데이터 파일을 선택해주세요.",
                                            font=("Consolas", 20))
            self.status_label.pack(pady=6)

            # 저장 경로 표시 레이블
            self.saved_path_label = ctk.CTkLabel(self.root, text="",
                                                font=("Consolas", 17),
                                                text_color="#1E88E5",
                                                wraplength=680)
            self.saved_path_label.pack(pady=0)

            # 진행 표시기
            self.progress = ctk.CTkProgressBar(self.root, mode='indeterminate',
                                              width=550, height=28)
            self.progress.set(0)

            # 백래쉬 계산 방법 버튼 (좌측 하단)
            info_btn_frame = ctk.CTkFrame(self.root, fg_color="transparent")
            info_btn_frame.pack(side="bottom", anchor="w", padx=20, pady=10)

            info_btn = ctk.CTkButton(info_btn_frame, text="ℹ️ 백래쉬 계산 방법",
                                    command=self.show_info,
                                    font=("돋움", 12, "bold"),
                                    fg_color="transparent",
                                    hover_color="#E3F2FD",
                                    text_color="#2196F3",
                                    border_width=2,
                                    border_color="#2196F3",
                                    width=140, height=35,
                                    corner_radius=8)
            info_btn.pack()

        def browse_file(self):
            """파일 선택 대화상자"""
            filename = filedialog.askopenfilename(
                title="데이터 파일 선택",
                filetypes=[("Data files", "*.xlsx *.xls *.csv"), ("Excel files", "*.xlsx *.xls"), ("CSV files", "*.csv"), ("All files", "*.*")]
            )
            
            if filename:
                self.selected_file = filename
                self.file_entry.delete(0, "end")
                self.file_entry.insert(0, filename)
                self.analyze_btn.configure(state="normal")
                self.status_label.configure(text="분석 준비 완료. '분석 시작' 버튼을 클릭하세요.",
                                           text_color="#4CAF50")
        
        def start_analysis(self):
            """분석 시작"""
            if not self.selected_file:
                messagebox.showerror("오류", "파일을 선택해주세요.")
                return

            # 버튼 비활성화
            self.analyze_btn.configure(state="disabled")
            self.status_label.configure(text="분석 중...", text_color="#2196F3")
            self.progress.pack(pady=10)
            self.progress.start()

            # 별도 스레드에서 분석 실행
            thread = threading.Thread(target=self.run_analysis)
            thread.start()
        
        def run_analysis(self):
            """분석 실행 (별도 스레드)"""
            try:
                output_path = process_file(self.selected_file)
                self.output_file = output_path

                # 성공 메시지
                self.root.after(0, self.analysis_complete)
            except Exception as e:
                # 오류 메시지
                self.root.after(0, lambda: self.analysis_error(str(e)))
        
        def analysis_complete(self):
            """분석 완료"""
            self.progress.stop()
            self.progress.pack_forget()
            self.status_label.configure(text="분석 완료! 결과가 Result 폴더에 저장되었습니다.",
                                       text_color="#4CAF50")

            # 저장된 파일 경로 표시
            if self.output_file:
                abs_path = os.path.abspath(self.output_file)
                self.saved_path_label.configure(text=f"저장 경로: {abs_path}")

            self.analyze_btn.configure(state="normal")

            messagebox.showinfo("성공",
                              f"분석이 성공적으로 완료되었습니다!\n\n"
                              f"결과 파일: {self.output_file}")
        
        def analysis_error(self, error_msg):
            """분석 오류"""
            self.progress.stop()
            self.progress.pack_forget()
            self.status_label.configure(text="분석 실패", text_color="#F44336")
            self.analyze_btn.configure(state="normal")

            messagebox.showerror("오류", f"분석 중 오류가 발생했습니다:\n\n{error_msg}")

        def open_result_folder(self):
            """Result 폴더를 탐색기로 열기"""
            result_dir = 'Result'
            if not os.path.exists(result_dir):
                os.makedirs(result_dir)

            # Windows에서 폴더 열기
            abs_path = os.path.abspath(result_dir)
            os.startfile(abs_path)

        def show_info(self):
            """백래쉬 계산 방법 설명 팝업"""
            info_text = (
                "📊 백래쉬 계산 방법\n\n"
                "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
                "🔹 X = 3% 지점 (back1):\n"
                "   • 모든 교차점에서 선형 보간\n"
                "   • 최대값(max) 선택\n\n"
                "🔹 X = -3% 지점 (back2):\n"
                "   • 모든 교차점에서 선형 보간\n"
                "   • 최소값(min) 선택\n\n"
                "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
                "✅ 백래쉬 = |back1 - back2|"
            )

            messagebox.showinfo("계산 방법 안내", info_text)
    
    # GUI 실행
    root = ctk.CTk()
    app = HysteresisApp(root)
    root.mainloop()

if __name__ == '__main__':
    # 커맨드라인 인수가 있으면 CLI 모드, 없으면 GUI 모드
    if len(sys.argv) > 1:
        main_cli()
    else:
        main_gui()
